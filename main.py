import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET

mapeamento = [
    ["DESCRICAO", "xProd"],
    ["REFERENCIA", "cEAN"],
    ["UNIDADEENT", "uCom"],
    ["PRECOCUSTO", "vUnCom"],
    ["CODCEST", "CEST"],
    ["CFOP", "CFOP"],
    ["CODTRIBUT00", "CST"],
    ["CODNBM", "NCM"],
]

mapeamentoFornecedor = [
    ["NOME", "xNome"],
    ["NOMEFANTASIA", "xFant"],
    ["BAIRRO", "xBairro"],
    ["CEP", "CEP"],
    ["CGCCPF", "CNPJ"],
    ["ENDERECO", "xLgr"],
    ["ESTADO", "UF"],
    ["Cidade", "xMun"],
    ["FONE", "fone"],
    ["INSCEST", "IE"],
    ["NUMERO", "nro"],
]

matriz_total = []
matriz_total_fornecedor = []

def buscar_arquivo_xml():
    arquivos_xml = filedialog.askopenfilenames(filetypes=[("XML files", "*.xml")])
    return arquivos_xml

def selecionar_arquivo_excel():
    arquivo_excel = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    return arquivo_excel

def selecionar_arquivo_excel():
    root = tk.Tk()
    root.withdraw()
    arquivo_excel = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    return arquivo_excel

def ler_xml_e_gerar_matriz(arquivo_xml):
    # Definindo o namespace
    namespace = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

    tree = ET.parse(arquivo_xml)
    root = tree.getroot()

    # Inicializa a matriz
    matriz = []

    # Procura por todas as tags <prod>
    produtos = root.findall(".//nfe:det", namespaces=namespace)

    for prod in produtos:
        # Inicializa uma nova linha na matriz
        linha = []
        produto_repetido = False
        referencia = prod.find(f'.//nfe:cEAN', namespaces=namespace)
        for x in matriz_total:
            if referencia.text == x[1]:
                produto_repetido = True
                break

        if produto_repetido:
            break
        for cabecalho, tag_xml in mapeamento:
            # Procura a tag correspondente dentro do produto
            element = prod.find(f'.//nfe:{tag_xml}', namespaces=namespace)
            if element is not None and element.text is not None:
                # Adiciona o valor à linha
                if tag_xml == 'CST':
                    linha.append(element.text.zfill(3))
                else:
                    linha.append(element.text)
            else:
                # Se a tag não existe ou está vazia, adiciona uma string vazia
                linha.append("")

        # Adiciona a linha à matriz
        matriz.append(linha)

    return matriz

def ler_xml_e_gerar_matriz_fornecedor(arquivo_xml):
    # Definindo o namespace
    namespace = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

    tree = ET.parse(arquivo_xml)
    root = tree.getroot()

    # Inicializa a matriz
    matriz = []
    # Procura por todas as tags <prod>
    fornecedor = root.findall(".//nfe:emit", namespaces=namespace)

    for prod in fornecedor:
        # Inicializa uma nova linha na matriz
        linha = []

        fornecedor_repetido = False
        cnpj = prod.find(f'.//nfe:CNPJ', namespaces=namespace)
        for x in matriz_total_fornecedor:
            if cnpj.text == x[4]:
                fornecedor_repetido = True
                break

        if fornecedor_repetido:
            break

        for cabecalho, tag_xml in mapeamentoFornecedor:
            # Procura a tag correspondente dentro do produto
            element = prod.find(f'.//nfe:{tag_xml}', namespaces=namespace)
            if element is not None and element.text is not None:
                # Adiciona o valor à linha
                linha.append(element.text)
            else:
                # Se a tag não existe ou está vazia, adiciona uma string vazia
                linha.append("")

        # Adiciona a linha à matriz
        matriz.append(linha)

    return matriz


def processar_matriz_e_atualizar_excel(matriz, arquivo_excel, mapeamento):
    # Cria um novo Workbook do Excel
    wb = Workbook()
    ws = wb.active

    # Adiciona os cabeçalhos ao Excel
    for col_num, (novo_cabecalho, _) in enumerate(mapeamento, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = novo_cabecalho

    # Preenche as colunas com os valores da matriz
    for row_num, linha in enumerate(matriz, 2):
        for col_num, valor in enumerate(linha, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}{row_num}'] = valor

    # Salva o Workbook
    wb.save(arquivo_excel)
    print(f'Arquivo Excel salvo em: {arquivo_excel}')

def adicionar_fornecedores_aba_existente(matrizFornecedor, arquivo_excel, mapeamentoFornecedor):
    # Carrega o Workbook existente
    wb = load_workbook(arquivo_excel)

    # Cria uma nova planilha chamada "Fornecedores"
    ws = wb.create_sheet(title="Fornecedores")

    # Adiciona os cabeçalhos ao Excel
    for col_num, (novo_cabecalho, _) in enumerate(mapeamentoFornecedor, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = novo_cabecalho

    # Preenche as colunas com os valores da matriz
    for row_num, linha in enumerate(matrizFornecedor, 2):
        for col_num, valor in enumerate(linha, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}{row_num}'] = valor

    # Salva o Workbook com a nova aba
    wb.save(arquivo_excel)


def ajustar_tamanho_colunas(arquivo_excel):
    # Carrega o Workbook existente
    wb = load_workbook(arquivo_excel)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Itera sobre todas as colunas
        for col in ws.columns:
            max_length = 0
            # Itera sobre todas as células da coluna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    # Salva o Workbook com as colunas ajustadas
    wb.save(arquivo_excel)

def gerar_excel():
    arquivos_xml = buscar_arquivo_xml()
    if arquivos_xml:
        arquivo_excel = selecionar_arquivo_excel()
        if arquivo_excel:
            for arquivo_xml in arquivos_xml:
                matriz = ler_xml_e_gerar_matriz(arquivo_xml)
                matrizFornecedor = ler_xml_e_gerar_matriz_fornecedor(arquivo_xml)
                matriz_total.extend(matriz)
                matriz_total_fornecedor.extend(matrizFornecedor)

            processar_matriz_e_atualizar_excel(matriz_total, arquivo_excel, mapeamento)
            adicionar_fornecedores_aba_existente(matriz_total_fornecedor, arquivo_excel, mapeamentoFornecedor)
            ajustar_tamanho_colunas(arquivo_excel)
            mensagem = f"O Excel foi gerado com sucesso para o arquivo: {arquivo_excel}"
            messagebox.showinfo("Concluído", mensagem)

# Criando a interface gráfica
def criar_janela():
    janela = tk.Tk()
    janela.title("Gerador de Excel")
    janela.geometry("300x200")

    def fechar_janela():
        janela.destroy()

    # Botão para gerar o Excel
    botao_gerar_excel = tk.Button(janela, text="Gerar Excel", command=gerar_excel)
    botao_gerar_excel.pack(pady=20)

    janela.protocol("WM_DELETE_WINDOW", fechar_janela)

    janela.mainloop()

if __name__ == "__main__":
    criar_janela()
